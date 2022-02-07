"""
@author = 'YuanXu'
@date = '2022/01/21'
功能：三代靶向耐药流程--结果整理和美化
"""
import argparse
import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

def myParser():
    parser = argparse.ArgumentParser(description = 'AMR_TF module')
    parser.add_argument('-p', '--pc', help = "输入PC barcode号，例：barcode17" , default = 'close')
    parser.add_argument('-n', '--nc', help = "输入NC barcode号，例：barcode18" , default = 'close')
    args = parser.parse_args()
    return args

def judge_input_fastq(input_sample):
    input_sample = os.path.abspath(input_sample)
    dir_list = os.listdir(input_sample) #文件夹内部的情况
    input_fastq_list = []
    for i in dir_list:
        input_fastq_list.append(i)
    return input_fastq_list


def reset_col(filename):
    wb=load_workbook(filename)
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        df=pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)]=list(df.columns)    #把标题行附加到最后一行
        for col in df.columns:
            index=list(df.columns).index(col)   #列序号
            letter=get_column_letter(index+1)   #列字母
            collen=df[col].apply(lambda x:len(str(x).encode())).max()   #获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width=collen*1.2+4    #也就是列宽为最大长度*1.2 可以自己调整
    wb.save(filename)

def beautiful_excel(filename):
    reset_col(filename)
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]] #打开第一个sheet
    
    al = Alignment(horizontal = 'center' , vertical = 'center')
    pf = PatternFill(patternType='solid', fgColor = '87CEEB')
    header_font = Font(bold = True, name = 'Times New Roman', size = 12, color = '00000000')
    for col in range(1, ws.max_column + 1):
        ws.cell(1,col).alignment = al
        ws.cell(1,col).font = header_font
        ws.cell(1,col).fill = pf
        
    pf2 = PatternFill(patternType = 'solid', fgColor = 'FFFFFF')
    header_font2 = Font(bold = True, name = 'Times New Roman', size = 11, color = '00000000')
    for row in range(2, ws.max_row - 1):
        ws.cell(row,1).alignment = al
        ws.cell(row,1).fill = pf2
        ws.cell(row,1).font = header_font2
        
    pf3 = PatternFill(patternType = None , fgColor = 'FFFFFF')
    header_font3 = Font(name = 'Times New Roman', size = 11, color = '00000000')
    for col in range(2, ws.max_column + 1):
        for row in range(2, ws.max_row - 1):
            ws.cell(row,col).alignment = al
            ws.cell(row,col).fill = pf3
            ws.cell(row,col).font = header_font3
            
    pf4 = PatternFill(patternType = None , fgColor = 'FFFFFF')
    header_font4 = Font(bold = True, name = 'Arial', size = 12, color = '00000000')
    for row in range(ws.max_row - 1, ws.max_row + 1):
        ws.cell(row,1).alignment = al
        ws.cell(row,1).fill = pf4
        ws.cell(row,1).font = header_font4
        
    pf5 = PatternFill(patternType = None , fgColor = 'FFFFFF')
    header_font5 = Font(bold = True, name = 'Arial', size = 11, color = '00000000')
    for col in range(2, ws.max_column + 1):
        for row in range(ws.max_row - 1, ws.max_row + 1):
            ws.cell(row,col).alignment = al
            ws.cell(row,col).fill = pf5
            ws.cell(row,col).font = header_font5
            
    wb.save(filename)

def AMR_TF(filename, PC_sample, NC_sample):
    df = pd.read_excel(filename, 'AMR_result', index_col = 0 ).fillna('-')
    #PC_sample = PC_sample.replace('barcode', 'BC')
    #NC_sample = NC_sample.replace('barcode', 'BC')
    df2 = df.rename(columns={PC_sample:'PC', NC_sample:'NC'})
    df2.drop(['Total', 'Total-phage'], inplace=True)
    
    barcode_list = list(df2.columns)
    barcode_list.remove('PC')
    barcode_list.remove('NC')
    
    for i in barcode_list:
        if df2.loc['phage', i] > 0.8*df2.loc['phage', 'PC']:
            bool1 = (df2[i] >= 100).tolist()
            bool2 = (df2[i] > 0.4*df2['PC']).tolist()
            bool3 = (df2[i] >= 2*df2['NC']).tolist()
            bool4 = (df2[i] >= 600).tolist()
            bool_result = []
            for j in range(len(bool1)): #未加长度相等控制判断
                bool_result.append(str(bool1[j] and bool2[j] and bool3[j] or bool4[j]))
                bool_result = [i.replace('True', 'T') for i in bool_result]
                bool_result = [i.replace('False', 'F') for i in bool_result]
            df2[i+'_bool'] = bool_result
            del df2[i]
    return df2

def beautiful_excel_TF(filename):
    reset_col(filename)
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]] #打开第一个sheet
    
    al = Alignment(horizontal = 'center' , vertical = 'center')
    pf = PatternFill(patternType='solid', fgColor = '87CEEB')
    header_font = Font(bold = True, name = 'Times New Roman', size = 12, color = '00000000')
    for col in range(1, ws.max_column + 1):
        ws.cell(1,col).alignment = al
        ws.cell(1,col).font = header_font
        ws.cell(1,col).fill = pf
        
    pf2 = PatternFill(patternType = 'solid', fgColor = 'FFFFFF')
    header_font2 = Font(bold = True, name = 'Times New Roman', size = 11, color = '00000000')
    for row in range(2, ws.max_row + 1):
        ws.cell(row,1).alignment = al
        ws.cell(row,1).fill = pf2
        ws.cell(row,1).font = header_font2
        
    pf3 = PatternFill(patternType = None , fgColor = 'FFFFFF')
    pf3_t = PatternFill(patternType = "solid" , fgColor = '00FFFF00')
    header_font3 = Font(name = 'Times New Roman', size = 11, color = '00000000')
    header_font3_t = Font(bold = True, name = 'Times New Roman', size = 11, color = '00000000')
    for col in range(2, ws.max_column + 1):
        for row in range(2, ws.max_row + 1):
            if ws.cell(row,col).value == 'T':
                ws.cell(row,col).alignment = al
                ws.cell(row,col).fill = pf3_t
                ws.cell(row,col).font = header_font3_t
            else:
                ws.cell(row,col).alignment = al
                ws.cell(row,col).fill = pf3
                ws.cell(row,col).font = header_font3
                
    wb.save(filename)

if __name__ == '__main__':
    args = myParser()
    '''
    判断真假阳性
    '''
    if args.pc != 'close' and args.pc != 'close':
        amr_tf_df = AMR_TF('result/AMR_result.xlsx', args.pc, args.nc)
        amr_tf_df.to_excel('result/AMR_TF_result.xlsx', sheet_name = 'AMR_TF_result', index=True) #写入到一个新excel表中
        beautiful_excel_TF('result/AMR_TF_result.xlsx')
