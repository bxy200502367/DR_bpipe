"""
@author = 'YuanXu'
@date = '2022/01/21'
功能：三代靶向耐药流程--结果整理和美化
"""
import argparse
import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

def myParser():
    parser = argparse.ArgumentParser(description = 'Merge Result')
    parser.add_argument('-i', '--input_fastq', help = "输入测序原始数据文件夹，fastq格式文件夹")
    args = parser.parse_args()
    return args

def judge_input_fastq(input_sample):
    input_sample = os.path.abspath(input_sample)
    dir_list = os.listdir(input_sample) #文件夹内部的情况
    input_fastq_list = []
    for i in dir_list:
        input_fastq_list.append(i)
    return input_fastq_list


def reset_col(filename):
    wb=load_workbook(filename)
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        df=pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)]=list(df.columns)    #把标题行附加到最后一行
        for col in df.columns:
            index=list(df.columns).index(col)   #列序号
            letter=get_column_letter(index+1)   #列字母
            collen=df[col].apply(lambda x:len(str(x).encode())).max()   #获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width=collen*1.2+4    #也就是列宽为最大长度*1.2 可以自己调整
    wb.save(filename)

def beautiful_excel(filename):
    reset_col(filename)
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]] #打开第一个sheet
    
    al = Alignment(horizontal = 'center' , vertical = 'center')
    pf = PatternFill(patternType='solid', fgColor = '87CEEB')
    header_font = Font(bold = True, name = 'Times New Roman', size = 12, color = '00000000')
    for col in range(1, ws.max_column + 1):
        ws.cell(1,col).alignment = al
        ws.cell(1,col).font = header_font
        ws.cell(1,col).fill = pf
        
    pf2 = PatternFill(patternType = 'solid', fgColor = 'FFFFFF')
    header_font2 = Font(bold = True, name = 'Times New Roman', size = 11, color = '00000000')
    for row in range(2, ws.max_row - 1):
        ws.cell(row,1).alignment = al
        ws.cell(row,1).fill = pf2
        ws.cell(row,1).font = header_font2
        
    pf3 = PatternFill(patternType = None , fgColor = 'FFFFFF')
    header_font3 = Font(name = 'Times New Roman', size = 11, color = '00000000')
    for col in range(2, ws.max_column + 1):
        for row in range(2, ws.max_row - 1):
            ws.cell(row,col).alignment = al
            ws.cell(row,col).fill = pf3
            ws.cell(row,col).font = header_font3
            
    pf4 = PatternFill(patternType = None , fgColor = 'FFFFFF')
    header_font4 = Font(bold = True, name = 'Arial', size = 12, color = '00000000')
    for row in range(ws.max_row - 1, ws.max_row + 1):
        ws.cell(row,1).alignment = al
        ws.cell(row,1).fill = pf4
        ws.cell(row,1).font = header_font4
        
    pf5 = PatternFill(patternType = None , fgColor = 'FFFFFF')
    header_font5 = Font(bold = True, name = 'Arial', size = 11, color = '00000000')
    for col in range(2, ws.max_column + 1):
        for row in range(ws.max_row - 1, ws.max_row + 1):
            ws.cell(row,col).alignment = al
            ws.cell(row,col).fill = pf5
            ws.cell(row,col).font = header_font5
            
    wb.save(filename)

def AMR_TF(filename, PC_sample, NC_sample):
    df = pd.read_excel(filename, 'AMR_result', index_col = 0 ).fillna('-')
    PC_sample = PC_sample.replace('barcode', 'BC')
    NC_sample = NC_sample.replace('barcode', 'BC')
    df2 = df.rename(columns={PC_sample:'PC', NC_sample:'NC'})
    df2.drop(['Total', 'Total-phage'], inplace=True)
    
    barcode_list = list(df2.columns)
    barcode_list.remove('PC')
    barcode_list.remove('NC')
    
    for i in barcode_list:
        if df2.loc['phage', i] > 0.8*df2.loc['phage', 'PC']:
            bool1 = (df2[i] >= 100).tolist()
            bool2 = (df2[i] > 0.4*df2['PC']).tolist()
            bool3 = (df2[i] >= 2*df2['NC']).tolist()
            bool4 = (df2[i] >= 600).tolist()
            bool_result = []
            for j in range(len(bool1)): #未加长度相等控制判断
                bool_result.append(str(bool1[j] and bool2[j] and bool3[j] or bool4[j]))
                bool_result = [i.replace('True', 'T') for i in bool_result]
                bool_result = [i.replace('False', 'F') for i in bool_result]
            df2[i+'_bool'] = bool_result
            del df2[i]
    return df2

def beautiful_excel_TF(filename):
    reset_col(filename)
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]] #打开第一个sheet
    
    al = Alignment(horizontal = 'center' , vertical = 'center')
    pf = PatternFill(patternType='solid', fgColor = '87CEEB')
    header_font = Font(bold = True, name = 'Times New Roman', size = 12, color = '00000000')
    for col in range(1, ws.max_column + 1):
        ws.cell(1,col).alignment = al
        ws.cell(1,col).font = header_font
        ws.cell(1,col).fill = pf
        
    pf2 = PatternFill(patternType = 'solid', fgColor = 'FFFFFF')
    header_font2 = Font(bold = True, name = 'Times New Roman', size = 11, color = '00000000')
    for row in range(2, ws.max_row + 1):
        ws.cell(row,1).alignment = al
        ws.cell(row,1).fill = pf2
        ws.cell(row,1).font = header_font2
        
    pf3 = PatternFill(patternType = None , fgColor = 'FFFFFF')
    pf3_t = PatternFill(patternType = "solid" , fgColor = '00FFFF00')
    header_font3 = Font(name = 'Times New Roman', size = 11, color = '00000000')
    header_font3_t = Font(bold = True, name = 'Times New Roman', size = 11, color = '00000000')
    for col in range(2, ws.max_column + 1):
        for row in range(2, ws.max_row + 1):
            if ws.cell(row,col).value == 'T':
                ws.cell(row,col).alignment = al
                ws.cell(row,col).fill = pf3_t
                ws.cell(row,col).font = header_font3_t
            else:
                ws.cell(row,col).alignment = al
                ws.cell(row,col).fill = pf3
                ws.cell(row,col).font = header_font3
                
    wb.save(filename)

def beautiful_excel_SNP(filename):
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]] #打开第一个sheet
    
    al = Alignment(horizontal = 'center' , vertical = 'center')
    pf = PatternFill(patternType='solid', fgColor = '87CEEB')
    header_font = Font(bold = True, name = 'Times New Roman', size = 12, color = '00000000')
    for col in range(1, ws.max_column + 1):
        ws.cell(1,col).alignment = al
        ws.cell(1,col).font = header_font
        ws.cell(1,col).fill = pf
        
    pf2 = PatternFill(patternType = 'solid', fgColor = 'FFFFFF')
    header_font2 = Font(bold = True, name = 'Times New Roman', size = 11, color = '00000000')
    for row in range(2, ws.max_row + 1):
        ws.cell(row,1).alignment = al
        ws.cell(row,1).fill = pf2
        ws.cell(row,1).font = header_font2
        
    pf3 = PatternFill(patternType = None , fgColor = 'FFFFFF')
    pf3_t = PatternFill(patternType = "solid" , fgColor = '00FFFF00')
    header_font3 = Font(name = 'Times New Roman', size = 11, color = '00000000')
    header_font3_t = Font(bold = True, name = 'Times New Roman', size = 11, color = '00000000')
    for col in range(2, ws.max_column + 1):
        for row in range(2, ws.max_row + 1):
            ws.cell(row,col).alignment = al
            ws.cell(row,col).fill = pf3
            ws.cell(row,col).font = header_font3
            
    wb.save(filename)

if __name__ == '__main__':
    args = myParser()
    input_fastq_file = judge_input_fastq(args.input_fastq) #获得barcode列表
    
    '''
    整理AMR结果
    '''
    AMR_result_list = [f"{i}_analysis/{i}_AMR_analysis/{i}.AMR.result_detail.txt" for i in input_fastq_file]
    new_list = []
    new_list_detail = []
    for i in AMR_result_list:
        dataframe = pd.read_csv(i, index_col = 0, sep = '\t')
        dataframe_copy = dataframe.copy(deep = True)
        new_list_detail.append(dataframe_copy)
        del dataframe['Ave_Length']
        del dataframe['Ave_Score']
        del dataframe['Ave_Identity']
        new_list.append(dataframe)
        
    df = pd.concat(new_list, axis=1) #多个DataFrame合并为一个
    df.loc['Total'] = df.apply(lambda x : x.sum()) #计算sum
    df.loc['Total-phage'] = df.loc['Total'] - df.loc['phage']
    #df_column_names = [i.replace('barcode', 'BC') for i in df.columns]
    #df.columns = df_column_names #更改barcode为BC，缩减
    
    df_detail = pd.concat(new_list_detail, axis=1)
    df.to_excel('result/AMR_result.xlsx', sheet_name = 'AMR_result', index=True) #写入到一个新excel表中
    df_detail.to_excel('result/AMR_result_detail.xlsx', sheet_name = 'AMR_result_detail', index=True) #写入到一个新excel表中
    '''
    美化excel结果
    '''
    beautiful_excel('result/AMR_result.xlsx')
    beautiful_excel_TF('result/AMR_result_detail.xlsx')
    '''
    判断真假阳性
    '''
    
    '''
    整理SNP结果
    '''
    new_list_SNP = []
    for i in input_fastq_file:
        dataframe = pd.read_csv(f"{i}_analysis/{i}_SNP_analysis/{i}.SNP.result.txt", index_col = 0, sep = '\t', keep_default_na=False)
        data_types_dict = {i : str, 'Depth' : str}#转换数据格式
        dataframe2 = dataframe.astype(data_types_dict)
        dataframe2[i] = dataframe2[i] + "(" +dataframe2['Depth'] + ")"
        del dataframe2['Depth']
        new_list_SNP.append(dataframe2)
        
    df_SNP_result = pd.concat(new_list_SNP, axis=1) #多个DataFrame合并为一个
    df_SNP_result.to_excel('result/SNP_result.xlsx', sheet_name = 'SNP_result', index = True) #写入到一个新excel表中
    
    '''
    美化excel结果
    '''
    beautiful_excel_SNP('result/SNP_result.xlsx')
    
    '''
    整理SGA_001结果
    '''
    dataframe_SGA_001 = pd.read_csv("SGA_001_result.txt", index_col = 0, sep = '\t')
    dataframe_SGA_001.to_excel('result/SGA_001_result.xlsx', sheet_name = 'SGA_001_result', index = True) #写入到一个新excel表中
    beautiful_excel_TF('result/SGA_001_result.xlsx')
    
    '''
    整理SGA_002结果
    '''
    dataframe_SGA_002 = pd.read_csv("SGA_002_result.txt", index_col = 0, sep = '\t')
    dataframe_SGA_002.to_excel('result/SGA_002_result.xlsx', sheet_name = 'SGA_002_result', index = True) #写入到一个新excel表中
    beautiful_excel_TF('result/SGA_002_result.xlsx')
    
    '''
    整理SGA_003结果
    '''
    new_list_SGA_003 = []
    for i in input_fastq_file:
        dataframe_SGA_003 = pd.read_csv(f"{i}_analysis/{i}_SGA_003_analysis/{i}.SGA_003.result.txt", index_col = 0, sep = '\t', keep_default_na=False)
        data_types_dict = {i : str, 'Depth' : str}#转换数据格式
        dataframe2 = dataframe_SGA_003.astype(data_types_dict)
        dataframe2[i] = dataframe2[i] + " depth(" +dataframe2['Depth'] + ")"
        del dataframe2['Depth']
        new_list_SGA_003.append(dataframe2)
        
    df_SGA_003_result = pd.concat(new_list_SGA_003, axis=1) #多个DataFrame合并为一个
    df_SGA_003_result.to_excel('result/SGA_003_result.xlsx', sheet_name = 'SGA_003_result', index = True) #写入到一个新excel表中
    beautiful_excel_TF('result/SGA_003_result.xlsx')
    
    '''
    整理MLST结果
    '''
    new_list_MLST = []
    for i in input_fastq_file:
        dataframe_MLST = pd.read_csv(f"{i}_analysis/{i}_MLST_analysis/{i}.MLST.result.txt", index_col = 0, sep = '\t', keep_default_na=False)
        data_types_dict = {i : str, 'Depth' : str}#转换数据格式
        dataframe2 = dataframe_MLST.astype(data_types_dict)
        dataframe2[i] = dataframe2[i] + " depth(" +dataframe2['Depth'] + ")"
        del dataframe2['Depth']
        new_list_MLST.append(dataframe2)
        
    df_MLST_result = pd.concat(new_list_MLST, axis=1) #多个DataFrame合并为一个
    df_MLST_result.to_excel('result/MLST_result.xlsx', sheet_name = 'MLST_result', index = True) #写入到一个新excel表中
    beautiful_excel_TF('result/MLST_result.xlsx')

